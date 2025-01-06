from functools import cached_property
from itertools import chain
import sys
from typing import (
    Any,
    ClassVar,
    Generator,
    Iterable,
    List,
    Literal,
    MutableSet,
    Optional,
    Self,
    Tuple,
)
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from dataclasses import dataclass, field

from pyxlmapper.formatters import PrettyFormatter
from pyxlmapper.util import (
    camel_to_snake,
    class_name_from_str,
    dict_path_set,
    normalize,
    unwrap,
)


@dataclass
class InternalConfig:
    """
    Main config container class
    """

    raw_name: str
    """
    Internal node name
    """

    offset: Tuple[int, int]
    """
    Offset (rows, headers). Default (0, 0)
    """

    input_name: str = ""
    """
    The normalized name of spreadsheet header cell
    """

    output_name: str = ""
    """
    The name of the JSON property
    """

    optional: bool = False
    """
    Optional nodes might be skipped by the mapper if a spreadsheet does not have it
    """

    skip: bool = False
    """
    Skipped nodes are unconditionally ifnored
    """

    # Meta
    _overrides: MutableSet[str] = field(default_factory=set)
    """
    List of fields that was explicitly set by the config
    (for reflection purposes)
    """
    _own_fields: ClassVar[Iterable[str]] = (
        "offset",
        "input_name",
        "output_name",
        "optional",
        "skip",
    )
    """
    List of fields that config could have (no internal or meta fields)
    """

    @classmethod
    def from_config(cls, nodeconf: type):
        """
        nodedef is not really 'NodeConfig', but it has matching properties
        """
        name = nodeconf.__name__
        # Set fields
        input_name = getattr(nodeconf, "input_name", name)
        output_name = getattr(nodeconf, "output_name", camel_to_snake(name))
        offset = getattr(nodeconf, "offset", (0, 0))
        optional = getattr(nodeconf, "optional", False)
        skip = getattr(nodeconf, "optional", False)

        # Set metadata
        overrides = set([d for d in cls._own_fields if hasattr(nodeconf, d)])

        return cls(
            raw_name=name,
            offset=offset,
            input_name=input_name,
            output_name=output_name,
            optional=optional,
            skip=skip,
            _overrides=overrides,
        )

    @classmethod
    def from_input_name(cls, input_name: str):
        """
        Build config for given header name
        """
        raw_name = class_name_from_str(input_name)
        output_name = camel_to_snake(raw_name)

        overrides = set()

        if raw_name != input_name:
            overrides.add("input_name")

        return cls(
            raw_name=raw_name,
            output_name=output_name,
            input_name=input_name,
            offset=(0, 0),
            _overrides=overrides,
        )

    def __str__(self) -> str:
        return f"{self.raw_name}({self.input_name} -> {self.output_name})"

    def __eq__(self, value: object, /) -> bool:
        if not isinstance(value, InternalConfig):
            return False

        return (
            self.raw_name == value.raw_name
            and self.input_name == value.input_name
            and self.output_name == value.output_name
        )


@dataclass
class MapperNode:
    """
    Bi-directional tree
    """

    config: InternalConfig
    parent: Optional[Self] = None
    children: List[Self] = field(default_factory=list)

    @classmethod
    def from_config(cls, config: type):
        internal_config = InternalConfig.from_config(config)
        return cls(config=internal_config)

    @classmethod
    def infer(cls, internal_config: InternalConfig):
        return cls(config=internal_config)

    @cached_property
    def column_pos(self):
        if self.parent is None:
            return self.config.offset[1] + 1

        def index_of(lst, item):
            for idx, d in enumerate(lst):
                if d == item:
                    return idx
            return -1

        pos = index_of(self.parent.children, self)

        if pos > 0:
            sibling = self.parent.children[pos - 1]
        else:
            sibling = None

        previous_node_pos = (
            sibling.last.abs_pos[1]
            if sibling is not None
            else self.parent.abs_pos[1] - 1
        )

        return previous_node_pos + self.config.offset[1] + 1

    @cached_property
    def abs_pos(self) -> Tuple[int, int]:
        """
        Absolute position in spreadsheet coordinate space
        """
        if self.parent is None:
            return (self.config.offset[0], self.column_pos)

        return (
            self.parent.abs_pos[0] + self.config.offset[0] + 1,
            self.column_pos,
        )

    @property
    def is_root(self):
        return self.parent is None

    @property
    def is_leaf(self):
        return len(self.children) == 0

    @property
    def last(self):
        """
        rightmost child node
        """
        if len(self.children) > 0:
            return self.children[-1].last
        else:
            return self

    @property
    def root(self):
        node = self
        while node.parent is not None:
            node = node.parent
        return node

    @property
    def qualified_name(self):
        path = self.get_path()
        names = map(lambda d: d.config.output_name, path[1:])
        return ".".join(names)

    @property
    def cardinality(self):
        return len(self.children) + sum(d.cardinality for d in self.children)

    @property
    def coordinate(self):
        col = get_column_letter(self.abs_pos[1])
        if self.is_root:
            return "N/A"
        return f"{col}{self.abs_pos[0]}"

    def clear_caches(self):
        if hasattr(self, "abs_pos"):
            del self.abs_pos

        if hasattr(self, "column_pos"):
            del self.column_pos

    def add_child(self, child: Self):
        self.children.append(child)
        child.parent = self

    def get_path(self):
        path: List[Self] = []
        node = self
        while node is not None:
            path.append(node)
            node = node.parent
        return list(reversed(path))

    def get_leaves(self):
        return self._get_leaves([])

    def _get_leaves(self, leaves: List[Self]):
        if self.is_leaf:
            return leaves + [self]
        else:
            for child in self.children:
                leaves = child._get_leaves(leaves)
        return leaves

    def __iter__(self) -> Generator[Self, Any, None]:
        yield self
        for d in chain(*map(iter, self.children)):
            yield d

    def __repr__(self) -> str:
        return str(PrettyFormatter(self))

    def __str__(self) -> str:
        if self.is_root:
            return f"Node<{self.config.raw_name} ({self.coordinate}) -- ROOT>"
        return f"Node<{self.config.raw_name} ({self.coordinate}) '{self.config.input_name}' -> '{self.config.output_name}'>"


def read_classdef(cls: type, parent: MapperNode):
    for key, attr in cls.__dict__.items():
        if key.startswith("__"):
            continue

        if isinstance(attr, type):
            node = MapperNode.from_config(attr)
            parent.add_child(node)
            read_classdef(attr, node)


class SpreadsheetParserMeta(type):
    pass


class SpreadsheetMapper(metaclass=SpreadsheetParserMeta):
    """
    Base class for any mapper definintion
    """

    root: MapperNode

    def __new__(cls) -> Self:
        cls.root = MapperNode.from_config(cls)
        read_classdef(cls, cls.root)

        return super().__new__(cls)

    @classmethod
    def from_node(cls, node: MapperNode):
        self = cls()
        self.root = node
        return self

    def map_rows(self, ws: Worksheet, start_at: int):
        self._verify_augment(ws)

        nodes = self.root.get_leaves()

        row_index = start_at
        while True:
            # Receiving obj
            obj = {}

            for col_idx, node in enumerate(nodes):
                cell = ws.cell(row_index, node.column_pos)
                value = normalize(cell.value)

                # Stop on first blank row (cell)
                if col_idx == 0:
                    if value is None:
                        print("Finished")
                        return

                str_path = list(
                    map(lambda d: d.config.output_name, node.get_path()[1:])
                )
                dict_path_set(obj, str_path, value)

            yield obj
            sys.stdout.write(".")

            row_index += 1

    def _verify_augment(self: Self, ws: Worksheet):
        """
        Check if the mapper matches with actual spreadsheet header
        Checks presense and marks 'optional' nodes.
        If optional node is present, mapper would be modified to reflect that
        """
        tree = self.root
        # If set, node would drop any caches
        clear_caches_flag = False

        # Access from root to leaf, depth first
        for node in tree:
            if node.is_root:
                continue

            if clear_caches_flag:
                node.clear_caches()

            cell = ws.cell(*node.abs_pos)
            # Checl cell name actually matches name in the config

            if (expected := node.config.input_name) != (
                actual := normalize(unwrap(cell).value)
            ):
                # Optional nodes might be skipped
                if node.config.optional:
                    # Effectively remove node from the tree
                    if node.parent is not None:
                        node.parent.children.remove(node)
                        node.parent = None
                        clear_caches_flag = True
                        continue

                # NOTE this might be not as simple, but works as expected with known data sets
                # when parent is None pyxlmapper will pick node to the left
                if actual is None:
                    continue

                raise ValueError(
                    "Mapper config does not match to actual .xlsx file! "
                    f"Node '{node.qualified_name}' at ({node.coordinate}) "
                    f"was expected to have '{expected}', but actual header is '{actual}'. "
                    "Or mark node with `optional = True` or `skip = True`"
                )


def collapse_levels(levels: List[str]) -> List[str]:
    """
    Remove duplicate levels
    """
    collapsed = []
    for level in levels:
        head = collapsed[-1] if len(collapsed) > 0 else None
        if head != level:
            collapsed.append(level)

    return collapsed


def read_header(
    ws: Worksheet,
    height: int,
    width: int | Literal["auto"] | None = None,
    offset: Tuple[int, int] = (0, 0),
):
    if width is None:
        width = "auto"

    col_index = 1

    while True:
        levels = []
        for level in range(1, height + 1):
            cell = unwrap(ws.cell(level + offset[0], col_index + offset[1]))
            value = normalize(cell.value)
            if value is None:
                continue
            levels.append(value)

        # Remove duplicates
        levels = collapse_levels(levels)
        yield levels

        # Next cell
        col_index += 1

        # Done on max width reached
        if width != "auto":
            if col_index > width:
                break

        # Done on no data
        if len(levels) == 0:
            break


def merge(reciever: MapperNode, other: MapperNode) -> Tuple[MapperNode | None, int]:
    """
    :param other: linear tree (no more than 1 child per node)
    :returns: the first divergent node and carry-on offset (for the next column)
    """
    assert len(other.children) <= 1  # not tested for > 1

    # Empty root node
    if reciever.is_root and len(reciever.children) == 0:
        reciever.add_child(other)
        return None, 0

    # The last node which was merged
    first_divergent = other
    last_child = reciever.children[-1]
    if last_child.config == other.config:
        # Convertion
        if len(other.children) > 0:
            other_last_child = other.children[-1]
            first_divergent = other_last_child
            return merge(last_child, other_last_child)
        else:
            return first_divergent, 1
    else:
        # Divertion
        reciever.add_child(other)
        return first_divergent, 0


def infer(
    ws: Worksheet,
    height: int,
    name: str,
    width: int | Literal["auto"] | None = None,
    offset: Tuple[int, int] = (0, 0),
):
    """
    Read a header from a real Worksheet
    """
    header = read_header(ws, height, width, offset)
    root_config = InternalConfig(raw_name=name, offset=offset)

    # Handle non-empty root offset
    if offset != (0, 0):
        root_config._overrides.add("offset")

    root_node = MapperNode(config=root_config)

    contextual_offset = 0

    for levels in header:
        configs = [InternalConfig.from_input_name(d) for d in levels]
        parent = None
        for _, config in enumerate(configs):
            node = MapperNode.infer(internal_config=config)

            # Link nodes
            if parent is not None:
                parent.add_child(node)
            parent = node

        if parent is not None:
            # Offset should be carried to the next item
            first_divergent, carry_offset = merge(root_node, parent.root)
            if carry_offset > 0:
                contextual_offset += carry_offset
            if contextual_offset > 0 and carry_offset == 0:
                ofst = root_node.last.config.offset
                first_divergent.config.offset = (ofst[0], ofst[1] + contextual_offset)
                # Set metadata (for to str impl)
                first_divergent.config._overrides.add("offset")
                contextual_offset = 0

    return SpreadsheetMapper.from_node(root_node)
